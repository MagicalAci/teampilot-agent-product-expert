#!/usr/bin/env python3
"""Summarize workbook sheets and keyword coverage."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - import guard
    load_workbook = None
    IMPORT_ERROR = exc
else:
    IMPORT_ERROR = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Summarize an XLSX knowledge-base file.")
    parser.add_argument("file", help="Workbook path")
    parser.add_argument(
        "--keywords",
        default="",
        help="Comma-separated keywords to count across all cell text",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=2000,
        help="Maximum rows to inspect per sheet",
    )
    return parser.parse_args()


def summarize_sheet(sheet, keywords: list[str], max_rows: int) -> dict:
    keyword_counts = Counter()
    non_empty_rows = 0
    sample_headers = []

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_index > max_rows:
            break

        values = [str(cell).strip() for cell in row if cell not in (None, "")]
        if not values:
            continue

        non_empty_rows += 1
        if row_index == 1:
            sample_headers = values[:10]

        merged = " ".join(values)
        for keyword in keywords:
            keyword_counts[keyword] += merged.count(keyword)

    return {
        "sheet": sheet.title,
        "non_empty_rows": non_empty_rows,
        "sample_headers": sample_headers,
        "keyword_counts": dict(keyword_counts),
    }


def main() -> int:
    args = parse_args()
    if load_workbook is None:
        raise SystemExit(
            "openpyxl is required to read XLSX files. "
            "Run `python scripts/bootstrap_product_planning_tools.py repair` "
            "or `bash scripts/bootstrap-macos.sh repair` first. "
            f"Import failed: {IMPORT_ERROR}"
        )

    keywords = [item.strip() for item in args.keywords.split(",") if item.strip()]
    workbook = load_workbook(Path(args.file), read_only=True, data_only=True)
    result = {
        "file": str(Path(args.file)),
        "sheet_names": workbook.sheetnames,
        "sheets": [
            summarize_sheet(sheet, keywords, args.max_rows) for sheet in workbook.worksheets
        ],
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
